"""
x_follow_analyzer.py — the desktop triage desk.

Same product as the web dashboard in web/, same visual language, same shared
storage. The queue is the work: one account fills the card, you decide, it leaves,
the counter drops. Totals and ratios are reference material and live in the rail
where they cannot compete with the decision in front of you.

Design notes
------------
* The palette is the CSS palette, hand-mixed to flat hex because a Tk canvas has no
  alpha channel: pale eucalyptus desk, warm bond-paper card, and exactly one
  saturated colour (the stamp) which appears for a moment and leaves with the card.
* The stamp is the app's icon — a tilted frame with a check inside, drawn as pure
  geometry. Rotated *text* in a Tk canvas is a shaping risk in Persian, and reusing
  the icon means the mark you click and the mark on your taskbar are one object.
* matplotlib is gone. It was 60 MB of dependency drawing two charts of two numbers
  each, and the direction demotes statistics to a thin rail — where a split bar
  says the same thing in 8 pixels. Fewer dependencies also means fewer strangers
  who clone this repo and hit an install error.

Behaviour notes
---------------
* Archive parsing runs on a worker thread and reports back through a Queue polled
  by the Tk loop. It used to block the UI thread with a bare root.update(), which
  froze the window solid on a multi-gigabyte archive.
* History goes through history_store.add/remove, which re-read the file before
  writing. The old code kept a set in memory and wrote it whole, so running the
  web dashboard at the same time meant whichever app saved last erased the other's
  work.
* Animations are tracked and cancellable. A stamp still in flight when a new
  archive loads would otherwise fire against the new queue and retire an account
  nobody looked at.
"""

from __future__ import annotations

import math
import queue as pyqueue
import threading
import tkinter as tk
import webbrowser
import zipfile
from tkinter import filedialog, font as tkfont, messagebox, ttk

import archive_parser
import history_store

# ==============================================================================
# Palette — the CSS custom properties, flattened.
# ==============================================================================
# `rule` is the hairline on paper, `rule_desk` the same hairline on the desk: one
# colour mixed over two backgrounds, which is what color-mix() does in the browser
# and what a canvas cannot do for us.
PALETTES = {
    "light": {
        "desk": "#BAC5BF",
        "desk_deep": "#A9B5AF",
        "card": "#FCFBF7",
        "ink": "#171C1A",
        "pencil": "#5A6560",
        "pine": "#1F4E46",
        "stamp": "#9E3324",
        "rule": "#DCDCD8",
        "rule_desk": "#A3ADA8",
        "shadow": "#98A5A0",
        "ghost": "#EFEFE9",
        "on_pine": "#FCFBF7",
        "field": "#F1F0EB",
        "row_hi": "#F2F1EC",
    },
    "dark": {
        "desk": "#171D1B",
        "desk_deep": "#101514",
        "card": "#232B28",
        "ink": "#E9EDE9",
        "pencil": "#98A39E",
        "pine": "#7FBFAC",
        "stamp": "#E0664F",
        "rule": "#3F4643",
        "rule_desk": "#343A38",
        "shadow": "#0B0F0E",
        "ghost": "#1D2422",
        "on_pine": "#101514",
        "field": "#1B211F",
        "row_hi": "#28312E",
    },
}

BATCH_SIZES = (5, 10, 20, 50)
ROW_CAP = 500          # the sheet is reference, not a database browser
CARD_MAX_W = 460
CARD_H = 248

UI_FONTS = ("Segoe UI", "Vazirmatn", "Noto Sans Arabic", "Tahoma", "DejaVu Sans")
DATA_FONTS = ("Cascadia Mono", "Consolas", "DejaVu Sans Mono", "Courier New")

T = {
    "title": "میز تریاژ فالو",
    "app": "میز تریاژ",
    "sub": "هیچ داده‌ای از این دستگاه بیرون نمی‌رود",
    "choose": "انتخاب آرشیو",
    "theme": "تغییر روشنایی",
    "ledger": "دفتر حساب",
    "following": "فالو می‌کنی",
    "followers": "فالوت می‌کنند",
    "mutual": "دوطرفه",
    "cleared": "رسیدگی‌شده",
    "one_way": "یک‌طرفه",
    "rate": "فالوبک گرفته‌ای",
    "skipped_one": "فایل نادیده‌گرفته‌شده",
    "skipped_why": "این فایل‌ها نام مشابه دارند ولی فهرست فالوور نیستند، پس شمرده نشدند.",
    "forget": "پاک کردن سابقه",
    "left": "مانده در صف",
    "v_queue": "صف",
    "v_all": "فهرست کامل",
    "v_done": "رسیدگی‌شده",
    "intro_eyebrow": "پروندهٔ جدید",
    "intro_lead": "آرشیو X را روی میز بگذار",
    "intro_note": "فایل zip همان چیزی است که از Settings ← Your account ← Download an archive می‌گیری. خواندنش روی همین دستگاه انجام می‌شود.",
    "busy_eyebrow": "در حال خواندن",
    "busy_lead": "آرشیو باز می‌شود…",
    "busy_note": "فقط فایل‌های فالوور و فالویینگ از حجم استخراج می‌شوند، پس آرشیو چند گیگابایتی هم سریع است.",
    "card_eyebrow": "فالو می‌کنی، فالوبک نداده",
    "card_id": "آیدی",
    "no_handle": "یوزرنیم ثبت نشده",
    "open_record": "باز کردن و ثبت",
    "skip": "بعدی",
    "empty_eyebrow": "صف تمام شد",
    "empty_lead": "همه را رسیدگی کردی",
    "empty_note": "می‌توانی رسیدگی‌شده‌ها را ببینی، یا آرشیو تازه‌ای بگذاری تا از اول شمرده شود.",
    "undo": "برگرداندن",
    "count": "تعداد",
    "find": "جست‌وجوی یوزرنیم یا آیدی",
    "th_handle": "یوزرنیم",
    "th_id": "آیدی",
    "excel": "خروجی اکسل",
    "keys": "Enter باز کردن و ثبت    ·    Space بعدی    ·    B دسته‌ای    ·    U برگرداندن",
    "sheet_hint": "برای باز کردن، روی سطر دوبار کلیک کن",
    "storage": "سابقه در پوشهٔ کاربری شما ذخیره می‌شود، نه در پوشهٔ برنامه:",
}


def _rot(points, cx, cy, deg):
    """
    Rotate (x, y) pairs about a centre.

    A Tk canvas has no transform matrix, so the tilt of the paper stack and of the
    stamp has to be computed by hand before the polygon is created.
    """
    rad = math.radians(deg)
    cos, sin = math.cos(rad), math.sin(rad)
    flat = []
    for x, y in points:
        dx, dy = x - cx, y - cy
        flat.extend((cx + dx * cos - dy * sin, cy + dx * sin + dy * cos))
    return flat


def _pick(root, candidates):
    """First installed family from a preference list, so one build works everywhere."""
    try:
        available = set(tkfont.families(root))
    except tk.TclError:
        return candidates[-1]
    for name in candidates:
        if name in available:
            return name
    return candidates[-1]


class Desk:
    # --------------------------------------------------------------------------
    # Construction
    # --------------------------------------------------------------------------
    def __init__(self, root):
        self.root = root
        self.theme = "light"
        self.view = "queue"

        # Analysis state, shaped exactly like the payload the web app renders so the
        # two frontends stay comparable.
        self.stats = {}
        self.one_way = []          # every account that owes a follow back
        self.queue = []            # the subset still to review, in working order
        self.done_ids = set()
        self.ignored = []
        self.username = ""

        self.undo_stack = []
        self.batch_size = 10
        self._anim = []            # scheduled after() ids, so a stamp can be cancelled
        self._busy_card = False    # input is ignored while a card is leaving
        self._parse_q = pyqueue.Queue()
        self._card_widgets = []    # embedded buttons, destroyed on every deck repaint
        self._deck_w = 0

        self.ui = _pick(root, UI_FONTS)
        self.data = _pick(root, DATA_FONTS)

        root.title(T["title"])
        root.geometry("1180x760")
        root.minsize(900, 640)

        self._build()
        self._bind_keys()
        self.apply_theme()

        # Whatever the web app recorded is already history; show it as such.
        self.done_ids = set(history_store.load())
        self._paint_figures()
        self._show_stage("intro")

    # --------------------------------------------------------------------------
    # Widget helpers
    # --------------------------------------------------------------------------
    def _btn(self, parent, text, command, kind="quiet", register=True):
        btn = tk.Button(
            parent, text=text, command=command, cursor="hand2",
            relief="flat", borderwidth=0, highlightthickness=0,
            padx=14, pady=7, font=(self.ui, 10, "bold"),
        )
        btn._kind = kind
        self._style_button(btn)
        # Buttons drawn onto the card are destroyed and rebuilt on every repaint, so
        # they must stay out of the theme sweep — configuring a destroyed widget is a
        # TclError, and the list would grow without bound as you work the queue.
        if register:
            self._buttons.append(btn)
        return btn

    def _style_button(self, btn):
        t = PALETTES[self.theme]
        looks = {
            "primary": (t["pine"], t["on_pine"], t["pine"], t["on_pine"]),
            "danger": (t["desk_deep"], t["stamp"], t["desk_deep"], t["stamp"]),
            "ghost": (t["desk"], t["pencil"], t["desk"], t["ink"]),
            "quiet": (t["card"], t["ink"], t["row_hi"], t["ink"]),
        }
        bg, fg, active_bg, active_fg = looks.get(btn._kind, looks["quiet"])
        btn.configure(bg=bg, fg=fg, activebackground=active_bg, activeforeground=active_fg)

    def _label(self, parent, text, role="ink", size=10, weight="normal", bg="desk", **kw):
        lbl = tk.Label(parent, text=text, font=(self.ui, size, weight), **kw)
        lbl._role, lbl._bg = role, bg
        self._labels.append(lbl)
        return lbl

    def _build(self):
        self._buttons = []
        self._labels = []
        self._panels = []          # (widget, palette key) for background sweeps

        # ---------------- masthead ----------------
        head = tk.Frame(self.root)
        head.pack(fill="x", padx=22, pady=(16, 0))
        self._panels.append((head, "desk"))

        # RTL: identity on the right, tools on the left.
        ident = tk.Frame(head)
        ident.pack(side="right")
        self._panels.append((ident, "desk"))

        mark = self._label(ident, "𝕏", size=20, weight="bold")
        mark.pack(side="right", padx=(0, 10))
        names = tk.Frame(ident)
        names.pack(side="right")
        self._panels.append((names, "desk"))
        self.lbl_app = self._label(names, T["app"], size=12, weight="bold", anchor="e")
        self.lbl_app.pack(fill="x")
        self._label(names, T["sub"], role="pencil", size=8, anchor="e").pack(fill="x")

        tools = tk.Frame(head)
        tools.pack(side="left")
        self._panels.append((tools, "desk"))
        self.btn_theme = self._btn(tools, "◐", self.toggle_theme, "icon")
        self.btn_theme.pack(side="left", padx=(0, 6))
        self._btn(tools, T["choose"], self.choose_archive, "quiet").pack(side="left", padx=(0, 6))
        self.lbl_who = self._label(tools, "", role="pencil", size=9)
        self.lbl_who.pack(side="left", padx=(0, 8))

        rule = tk.Frame(self.root, height=1)
        rule.pack(fill="x", pady=(14, 0))
        self._panels.append((rule, "rule_desk"))

        body = tk.Frame(self.root)
        body.pack(fill="both", expand=True, padx=22, pady=18)
        self._panels.append((body, "desk"))

        self._build_rail(body)
        self._build_stage(body)

    # ---------------- the rail ----------------
    def _build_rail(self, parent):
        rail = tk.Frame(parent, padx=14, pady=14, highlightthickness=1)
        rail.pack(side="right", fill="y", padx=(18, 0))
        rail.pack_propagate(False)
        rail.configure(width=232)
        self.rail = rail
        self._panels.append((rail, "desk_deep"))

        self._label(rail, T["ledger"], role="pencil", size=8, weight="bold",
                    bg="desk_deep", anchor="e").pack(fill="x", pady=(0, 8))

        self.figs = {}
        for key in ("following", "followers", "mutual", "cleared"):
            row = tk.Frame(rail)
            row.pack(fill="x", pady=1)
            self._panels.append((row, "desk_deep"))
            # Value on the left, name on the right: an RTL definition list.
            val = tk.Label(row, text="—", font=(self.data, 11, "bold"))
            val._role, val._bg = "ink", "desk_deep"
            self._labels.append(val)
            val.pack(side="left")
            self._label(row, T[key], role="pencil", size=9, bg="desk_deep").pack(side="right")
            self.figs[key] = val

        # The split bar: what the donut chart used to say, in eight pixels.
        # Every optional block gets a permanent container that is packed once, and the
        # block inside it is shown or hidden. Re-packing the block itself would append
        # it after its neighbours, so the rail would silently reorder as data arrived.
        self.split_box = tk.Frame(rail)
        self.split_box.pack(fill="x")
        self._panels.append((self.split_box, "desk_deep"))
        self.split = tk.Frame(self.split_box)
        self._panels.append((self.split, "desk_deep"))
        sep = tk.Frame(self.split, height=1)
        sep.pack(fill="x", pady=(12, 10))
        self._panels.append((sep, "rule_desk"))
        self.bar = tk.Canvas(self.split, height=8, highlightthickness=0, bd=0)
        self.bar.pack(fill="x")
        keys = tk.Frame(self.split)
        keys.pack(fill="x", pady=(6, 0))
        self._panels.append((keys, "desk_deep"))
        self.key_mutual = self._label(keys, "▪ " + T["mutual"], role="pine", size=8, bg="desk_deep")
        self.key_mutual.pack(side="right")
        self._label(keys, "▪ " + T["one_way"], role="pencil", size=8, bg="desk_deep").pack(side="right", padx=(0, 10))
        self.lbl_rate = self._label(self.split, "", role="pencil", size=9, bg="desk_deep", anchor="e")
        self.lbl_rate.pack(fill="x", pady=(6, 0))

        self.skipped_box = tk.Frame(rail)
        self.skipped_box.pack(fill="x")
        self._panels.append((self.skipped_box, "desk_deep"))
        self.skipped = tk.Frame(self.skipped_box)
        self._panels.append((self.skipped, "desk_deep"))
        sep2 = tk.Frame(self.skipped, height=1)
        sep2.pack(fill="x", pady=(12, 8))
        self._panels.append((sep2, "rule_desk"))
        self.lbl_skipped = self._label(self.skipped, "", role="pencil", size=8, bg="desk_deep",
                                       anchor="e", justify="right", wraplength=200)
        self.lbl_skipped.pack(fill="x")

        foot = tk.Frame(rail)
        foot.pack(side="bottom", fill="x")
        self._panels.append((foot, "desk_deep"))
        sep3 = tk.Frame(foot, height=1)
        sep3.pack(fill="x", pady=(0, 10))
        self._panels.append((sep3, "rule_desk"))
        self._label(foot, T["storage"], role="pencil", size=8, bg="desk_deep",
                    anchor="e", justify="right", wraplength=200).pack(fill="x")
        path = tk.Label(foot, text=str(history_store.history_path().parent),
                        font=(self.data, 7), anchor="e", justify="right", wraplength=200)
        path._role, path._bg = "pencil", "desk_deep"
        self._labels.append(path)
        path.pack(fill="x", pady=(2, 8))
        self.btn_forget = self._btn(foot, T["forget"], self.forget_all, "danger")
        self.btn_forget.pack(fill="x")
        self._btn(foot, T["excel"], self.export_excel, "quiet").pack(fill="x", pady=(6, 0))

    # ---------------- the stage ----------------
    def _build_stage(self, parent):
        stage = tk.Frame(parent)
        stage.pack(side="right", fill="both", expand=True)
        self._panels.append((stage, "desk"))

        # Five bands, packed once, top to bottom. Everything that appears and
        # disappears does so *inside* its band, so the vertical order of the stage is
        # fixed by construction instead of by a chain of pack(before=...) calls that
        # break the moment their anchor is hidden.
        bands = {}
        for name, opts in (("tally", {}), ("views", {}),
                           ("main", {"expand": True}), ("tray", {}), ("keys", {})):
            band = tk.Frame(stage)
            band.pack(fill="both" if opts.get("expand") else "x", **opts)
            self._panels.append((band, "desk"))
            bands[name] = band
        self.bands = bands

        # Signature element: the count that empties, and a spool that drains with it.
        self.tally = tk.Frame(bands["tally"])
        self._panels.append((self.tally, "desk"))
        self.lbl_count = tk.Label(self.tally, text="0", font=(self.data, 44, "bold"))
        self.lbl_count._role, self.lbl_count._bg = "ink", "desk"
        self._labels.append(self.lbl_count)
        self.lbl_count.pack()
        self._label(self.tally, T["left"], role="pencil", size=8, weight="bold").pack()
        self.spool = tk.Canvas(self.tally, height=4, width=340, highlightthickness=0, bd=0)
        self.spool.pack(pady=(8, 14))

        self.views = tk.Frame(bands["views"], padx=3, pady=3, highlightthickness=1)
        self._panels.append((self.views, "desk_deep"))
        self.view_btns = {}
        for key, label in (("queue", T["v_queue"]), ("all", T["v_all"]), ("done", T["v_done"])):
            b = tk.Button(self.views, text=label, cursor="hand2", relief="flat",
                          borderwidth=0, highlightthickness=0, padx=16, pady=4,
                          font=(self.ui, 10, "bold"),
                          command=lambda k=key: self.set_view(k))
            b.pack(side="right", padx=1)
            self.view_btns[key] = b

        # One canvas holds every state of the deck: intro, busy, live card, empty.
        # Drawing them instead of packing four frames is what makes the paper stack,
        # the tilt and the stamp possible at all.
        self.canvas = tk.Canvas(bands["main"], height=CARD_H + 44, highlightthickness=0, bd=0)
        self.canvas.pack(fill="x", pady=(16, 0))
        self.canvas.bind("<Configure>", self._on_resize)
        self._stage_mode = "intro"

        self.tray = tk.Frame(bands["tray"])
        self._panels.append((self.tray, "desk"))
        self.btn_undo = self._btn(self.tray, T["undo"], self.undo, "ghost")
        self.btn_undo.pack(side="left")
        self.btn_batch = self._btn(self.tray, "", self.open_batch, "quiet")
        self.btn_batch.pack(side="right")
        self.size_box = ttk.Combobox(self.tray, values=[str(n) for n in BATCH_SIZES],
                                     width=4, state="readonly", justify="center",
                                     font=(self.data, 10))
        self.size_box.set(str(self.batch_size))
        self.size_box.bind("<<ComboboxSelected>>", self._on_batch_size)
        self.size_box.pack(side="right", padx=(8, 8))
        self._label(self.tray, T["count"], role="pencil", size=8, weight="bold").pack(side="right")

        self.lbl_keys = self._label(bands["keys"], T["keys"], role="pencil", size=8)

        # ---------------- reference sheet ----------------
        self.sheet = tk.Frame(bands["main"], highlightthickness=1)
        self._panels.append((self.sheet, "card"))
        head = tk.Frame(self.sheet, padx=10, pady=8)
        head.pack(fill="x")
        self._panels.append((head, "card"))
        self.find = tk.Entry(head, font=(self.ui, 10), relief="flat", justify="right",
                             highlightthickness=1, bd=4)
        self.find.pack(side="right", fill="x", expand=True)
        self.find.bind("<KeyRelease>", lambda _e: self._paint_sheet())
        self.lbl_sheet_count = tk.Label(head, text="", font=(self.data, 9))
        self.lbl_sheet_count._role, self.lbl_sheet_count._bg = "pencil", "card"
        self._labels.append(self.lbl_sheet_count)
        self.lbl_sheet_count.pack(side="left", padx=(10, 0))
        sep = tk.Frame(self.sheet, height=1)
        sep.pack(fill="x")
        self._panels.append((sep, "rule"))

        wrap = tk.Frame(self.sheet)
        wrap.pack(fill="both", expand=True)
        self._panels.append((wrap, "card"))
        self.rows = ttk.Treeview(wrap, columns=("handle", "id"), show="headings",
                                 selectmode="browse", height=9)
        self.rows.heading("handle", text=T["th_handle"], anchor="e")
        self.rows.heading("id", text=T["th_id"], anchor="e")
        self.rows.column("handle", width=280, anchor="e")
        self.rows.column("id", width=200, anchor="e")
        scroll = ttk.Scrollbar(wrap, orient="vertical", command=self.rows.yview)
        self.rows.configure(yscrollcommand=scroll.set)
        scroll.pack(side="left", fill="y")
        self.rows.pack(side="right", fill="both", expand=True)
        self.rows.bind("<Double-1>", self._on_row_open)
        self.rows.bind("<Return>", self._on_row_open)
        self.lbl_more = tk.Label(self.sheet, text="", font=(self.ui, 8), anchor="e", padx=10, pady=6)
        self.lbl_more._role, self.lbl_more._bg = "pencil", "card"
        self._labels.append(self.lbl_more)

        # ---------------- notice strip ----------------
        # A bottom strip rather than a modal box: a blocked popup or an unreadable
        # file should not stop you working through the rest of the queue.
        self.notice = tk.Frame(self.root, padx=14, pady=10, highlightthickness=1)
        self.lbl_notice = tk.Label(self.notice, text="", font=(self.ui, 9), anchor="e", justify="right")
        self.lbl_notice._role, self.lbl_notice._bg = "ink", "card"
        self._labels.append(self.lbl_notice)
        self.lbl_notice.pack(side="right", fill="x", expand=True)
        close = tk.Button(self.notice, text="×", command=self.hush, cursor="hand2",
                          relief="flat", borderwidth=0, highlightthickness=0,
                          font=(self.ui, 12), padx=6, pady=0)
        close._kind = "ghost"
        self._buttons.append(close)
        close.pack(side="left")

    def _bind_keys(self):
        r = self.root
        r.bind("<Return>", lambda _e: self._key(self.open_and_record))
        r.bind("<space>", lambda _e: self._key(self.skip))
        r.bind("<Right>", lambda _e: self._key(self.skip))
        r.bind("<KeyPress-u>", lambda _e: self._key(self.undo))
        r.bind("<KeyPress-b>", lambda _e: self._key(self.open_batch))
        r.bind("<KeyPress-t>", lambda _e: self._key(self.toggle_theme))
        r.bind("<Control-o>", lambda _e: self.choose_archive())
        r.bind("<Control-z>", lambda _e: self.undo())
        r.bind("<Escape>", lambda _e: self.hush())
        # Through _key like the rest: searching for the account whose handle is "t3"
        # should type "t3", not flip the theme and jump to another view. Ctrl and Escape
        # above are safe unguarded — they are not characters anyone types into a field.
        for n, key in ((1, "queue"), (2, "all"), (3, "done")):
            r.bind(f"<KeyPress-{n}>", lambda _e, k=key: self._key(lambda: self.set_view(k)))

    def _key(self, action):
        """Shortcuts must not fire while you are typing in the search box."""
        if self.root.focus_get() in (self.find, self.size_box):
            return
        action()

    # --------------------------------------------------------------------------
    # Theme
    # --------------------------------------------------------------------------
    def toggle_theme(self):
        self.theme = "dark" if self.theme == "light" else "light"
        self.apply_theme()

    def apply_theme(self):
        t = PALETTES[self.theme]
        self.root.configure(bg=t["desk"])

        for widget, key in self._panels:
            try:
                widget.configure(bg=t[key])
            except tk.TclError:
                pass
        for widget, key in ((self.rail, "rule_desk"), (self.views, "rule_desk"),
                            (self.sheet, "rule"), (self.notice, "stamp")):
            try:
                widget.configure(highlightbackground=t[key], highlightcolor=t[key])
            except tk.TclError:
                pass

        for lbl in self._labels:
            lbl.configure(bg=t[lbl._bg], fg=t[lbl._role])

        for btn in self._buttons:
            self._style_button(btn)

        self.find.configure(bg=t["field"], fg=t["ink"], insertbackground=t["ink"],
                            highlightbackground=t["rule"], highlightcolor=t["pine"])

        style = ttk.Style()
        try:
            style.theme_use("clam")
        except tk.TclError:
            pass
        style.configure("Treeview", background=t["card"], fieldbackground=t["card"],
                        foreground=t["ink"], rowheight=30, borderwidth=0,
                        font=(self.data, 10))
        style.map("Treeview", background=[("selected", t["pine"])],
                  foreground=[("selected", t["on_pine"])])
        style.configure("Treeview.Heading", background=t["card"], foreground=t["pencil"],
                        font=(self.ui, 8, "bold"), borderwidth=0, padding=(8, 8))
        style.map("Treeview.Heading", background=[("active", t["row_hi"])])
        style.configure("Vertical.TScrollbar", background=t["desk_deep"],
                        troughcolor=t["card"], arrowcolor=t["pencil"], borderwidth=0)
        style.configure("TCombobox", fieldbackground=t["card"], background=t["card"],
                        foreground=t["ink"], arrowcolor=t["pencil"], borderwidth=0)
        self.rows.tag_configure("odd", background=t["row_hi"])
        self.rows.tag_configure("even", background=t["card"])
        self.rows.tag_configure("done", foreground=t["pencil"])

        self._paint_bar()
        self._paint_deck()

    # --------------------------------------------------------------------------
    # Painting
    # --------------------------------------------------------------------------
    def _show_stage(self, mode):
        """Swap what the stage shows: intro, busy, or the working queue."""
        self._stage_mode = mode
        working = mode == "work"
        if working:
            self.tally.pack()
            self.views.pack()
        else:
            self.tally.pack_forget()
            self.views.pack_forget()
        self.set_view("queue", repaint=False)
        self._paint_deck()

    def set_view(self, key, repaint=True):
        # Before an archive is loaded there is nothing to list, so the view keys are
        # inert. Without this, pressing 2 on the intro screen would leave view="all"
        # and the next load would open on the sheet instead of the queue.
        if self._stage_mode != "work" and key != "queue":
            return
        self.view = key
        working = self._stage_mode == "work"
        t = PALETTES[self.theme]
        for name, btn in self.view_btns.items():
            on = name == key and working
            btn.configure(bg=t["card"] if on else t["desk_deep"],
                          fg=t["ink"] if on else t["pencil"],
                          activebackground=t["card"] if on else t["desk_deep"],
                          activeforeground=t["ink"])

        on_queue = key == "queue"
        if on_queue or not working:
            self.sheet.pack_forget()
            self.canvas.pack(fill="x", pady=(16, 0))
        else:
            self.canvas.pack_forget()
            self.sheet.pack(fill="both", expand=True, pady=(16, 0))

        if working and on_queue:
            self.tray.pack(fill="x", pady=(16, 0))
            self.lbl_keys.pack(pady=(12, 0))
        else:
            self.tray.pack_forget()
            self.lbl_keys.pack_forget()

        if repaint:
            self._paint_deck()
            self._paint_sheet()
            self._paint_tally()

    def _paint_figures(self):
        s = self.stats
        fmt = lambda n: f"{n:,}" if isinstance(n, int) else "—"
        self.figs["following"].configure(text=fmt(s.get("following")))
        self.figs["followers"].configure(text=fmt(s.get("followers")))
        self.figs["mutual"].configure(text=fmt(s.get("mutuals")))
        self.figs["cleared"].configure(text=f"{len(self.done_ids):,}")

        if s:
            self.split.pack(fill="x")
            self.lbl_rate.configure(text=f'{s.get("win_rate", 0)}٪  {T["rate"]}')
        else:
            self.split.pack_forget()

        if self.ignored:
            self.skipped.pack(fill="x")
            names = "\n".join(self.ignored[:6])
            self.lbl_skipped.configure(
                text=f'{len(self.ignored)} {T["skipped_one"]}\n{T["skipped_why"]}\n{names}')
        else:
            self.skipped.pack_forget()

        self.lbl_who.configure(text=f"@{self.username}" if self.username else "")
        self.lbl_app.configure(text=T["app"])
        self._paint_bar()

    def _paint_bar(self):
        c = self.bar
        c.delete("all")
        t = PALETTES[self.theme]
        c.configure(bg=t["rule_desk"])
        s = self.stats
        total = s.get("following") or 0
        if not total:
            return
        width = max(c.winfo_width(), 200)
        mutual_w = int(width * (s.get("mutuals", 0) / total))
        # RTL: the mutual share grows from the right edge.
        c.create_rectangle(width - mutual_w, 0, width, 8, fill=t["pine"], width=0)
        c.create_rectangle(0, 0, width - mutual_w, 8, fill=t["pencil"], width=0)

    def _paint_tally(self):
        left = len(self.queue)
        self.lbl_count.configure(text=f"{left:,}")
        total = len(self.one_way) or 1
        c = self.spool
        c.delete("all")
        t = PALETTES[self.theme]
        c.configure(bg=t["rule_desk"])
        width = max(c.winfo_width(), 340)
        fill = int(width * left / total)
        if fill:
            c.create_rectangle(width - fill, 0, width, 4, fill=t["pine"], width=0)

        self.btn_batch.configure(text=f'باز کردن {self.batch_size} تای بعدی')
        self.btn_batch.configure(state="normal" if self.queue else "disabled")
        self.btn_undo.configure(state="normal" if self.undo_stack else "disabled")

    def _on_resize(self, event):
        # Configure fires continuously while dragging; only a real width change is
        # worth a full redraw of the deck.
        if abs(event.width - self._deck_w) < 3:
            return
        self._deck_w = event.width
        self._paint_deck()
        self._paint_bar()
        self._paint_tally()

    def _card_box(self):
        width = max(self.canvas.winfo_width(), 420)
        w = min(width - 40, CARD_MAX_W)
        x0 = (width - w) // 2
        return x0, 16, x0 + w, 16 + CARD_H

    def _paper(self, box, fill, outline, shadow=False):
        x0, y0, x1, y1 = box
        t = PALETTES[self.theme]
        if shadow:
            self.canvas.create_rectangle(x0 + 2, y0 + 5, x1 + 2, y1 + 5,
                                         fill=t["shadow"], width=0)
        return self.canvas.create_rectangle(x0, y0, x1, y1, fill=fill, outline=outline)

    def _paint_deck(self):
        c = self.canvas
        for w in self._card_widgets:
            w.destroy()
        self._card_widgets = []
        c.delete("all")
        t = PALETTES[self.theme]
        c.configure(bg=t["desk"])
        box = self._card_box()
        x0, y0, x1, y1 = box
        right = x1 - 24
        body = (x1 - x0) - 48          # the paper minus its margins, for wrapped text

        if self._stage_mode == "intro":
            self._dashed(box)
            self._text(right, y0 + 26, T["intro_eyebrow"], "pencil", 8, "bold")
            self._text(right, y0 + 52, T["intro_lead"], "ink", 15, "bold")
            self._text(right, y0 + 96, T["intro_note"], "pencil", 9, wrap=body)
            self._embed(self._btn(c, T["choose"], self.choose_archive, "primary", register=False),
                        right, y0 + 168, anchor="ne")
            return

        if self._stage_mode == "busy":
            self._dashed(box)
            self._text(right, y0 + 26, T["busy_eyebrow"], "pencil", 8, "bold")
            self._text(right, y0 + 52, T["busy_lead"], "ink", 15, "bold")
            self._text(right, y0 + 96, T["busy_note"], "pencil", 9, wrap=body)
            self._text(right, y0 + 168, "• • •", "pencil", 14, "bold")
            return

        if self.view != "queue":
            return

        if not self.queue:
            self._paper(box, t["card"], t["rule"], shadow=True)
            self._text(right, y0 + 26, T["empty_eyebrow"], "pencil", 8, "bold")
            self._text(right, y0 + 52, T["empty_lead"], "ink", 15, "bold")
            self._text(right, y0 + 96, T["empty_note"], "pencil", 9, wrap=body)
            self._embed(self._btn(c, T["v_done"], lambda: self.set_view("done"), "quiet", register=False),
                        right, y0 + 168, anchor="ne")
            return

        # Depth is literal: the stack behind the card is the work still to come.
        for depth, tilt in ((3, 0.35), (2, -0.6), (1, 0.5)):
            if len(self.queue) <= depth:
                continue
            dy = depth * 6
            pts = [(x0, y0 + dy), (x1, y0 + dy), (x1, y1 + dy), (x0, y1 + dy)]
            cx, cy = (x0 + x1) / 2, (y0 + y1) / 2 + dy
            c.create_polygon(_rot(pts, cx, cy, tilt), fill=t["ghost"], outline=t["rule"])

        self._paper(box, t["card"], t["rule"], shadow=True)
        person = self.queue[0]
        handle = person.get("username") or ""
        self._text(right, y0 + 26, T["card_eyebrow"], "pencil", 8, "bold")
        self._text(right, y0 + 56, f"@{handle}" if handle else T["no_handle"],
                   "ink", 19 if handle else 13, "bold", mono=bool(handle))
        self._text(right, y0 + 104, T["card_id"], "pencil", 8, "bold")
        self.canvas.create_text(right - 34, y0 + 104, text=person["account_id"], anchor="ne",
                                font=(self.data, 9), fill=t["ink"])

        self._embed(self._btn(c, T["open_record"], self.open_and_record, "primary", register=False),
                    right, y0 + 168, anchor="ne")
        self._embed(self._btn(c, T["skip"], self.skip, "quiet", register=False),
                    right - 132, y0 + 168, anchor="ne")

    def _dashed(self, box):
        x0, y0, x1, y1 = box
        t = PALETTES[self.theme]
        self.canvas.create_rectangle(x0, y0, x1, y1, outline=t["rule_desk"], dash=(4, 4))

    def _text(self, x, y, text, role, size, weight="normal", mono=False, wrap=None):
        # wrap= hands the line breaking to Tk. Breaking the paragraphs by hand would
        # look right in one font and overflow the card in the next, and the font we
        # actually get depends on what the machine has installed.
        t = PALETTES[self.theme]
        family = self.data if mono else self.ui
        opts = {} if wrap is None else {"width": max(int(wrap), 120)}
        return self.canvas.create_text(x, y, text=text, anchor="ne", justify="right",
                                       font=(family, size, weight), fill=t[role], **opts)

    def _embed(self, widget, x, y, anchor="ne"):
        self._card_widgets.append(widget)
        self.canvas.create_window(x, y, window=widget, anchor=anchor)
        return widget

    def _stamp_mark(self):
        """
        The app's icon, drawn on the card: a tilted frame with a check inside.

        Pure geometry on purpose. Rotated *text* in a Tk canvas risks breaking the
        joined letterforms of Persian, and reusing the icon means the mark on the
        card and the mark on the taskbar are the same object.
        """
        t = PALETTES[self.theme]
        x0, y0, x1, _ = self._card_box()
        cx, cy = x0 + 74, y0 + 96
        frame = [(cx - 52, cy - 21), (cx + 52, cy - 21), (cx + 52, cy + 21), (cx - 52, cy + 21)]
        items = [self.canvas.create_polygon(_rot(frame, cx, cy, -11), outline=t["stamp"],
                                            fill="", width=3)]
        tick = _rot([(cx - 20, cy + 1), (cx - 7, cy + 14), (cx + 20, cy - 15)], cx, cy, -11)
        items.append(self.canvas.create_line(*tick, fill=t["stamp"], width=3,
                                             capstyle="round", joinstyle="round"))
        return items

    def _paint_sheet(self):
        if self._stage_mode != "work" or self.view == "queue":
            return
        source = self.one_way if self.view == "all" else [
            p for p in self.one_way if p["account_id"] in self.done_ids]
        term = self.find.get().strip().lower()
        if term:
            source = [p for p in source
                      if term in (p.get("username") or "").lower() or term in p["account_id"]]

        self.rows.delete(*self.rows.get_children())
        for i, person in enumerate(source[:ROW_CAP]):
            done = person["account_id"] in self.done_ids
            tags = ["odd" if i % 2 else "even"]
            if done:
                tags.append("done")
            handle = person.get("username")
            self.rows.insert("", "end", iid=person["account_id"],
                             values=(f"@{handle}" if handle else "—", person["account_id"]),
                             tags=tuple(tags))

        self.lbl_sheet_count.configure(text=f"{len(source):,}")
        if len(source) > ROW_CAP:
            self.lbl_more.configure(
                text=f"{ROW_CAP:,} سطر اول نمایش داده شد. برای پیدا کردن بقیه جست‌وجو کن.")
            self.lbl_more.pack(fill="x")
        else:
            self.lbl_more.configure(text=T["sheet_hint"])
            self.lbl_more.pack(fill="x")

    # --------------------------------------------------------------------------
    # Notices
    # --------------------------------------------------------------------------
    def say(self, message):
        self.lbl_notice.configure(text=message)
        self.notice.pack(side="bottom", fill="x", padx=22, pady=(0, 14))

    def hush(self):
        self.notice.pack_forget()

    # --------------------------------------------------------------------------
    # Loading an archive
    # --------------------------------------------------------------------------
    def choose_archive(self):
        path = filedialog.askopenfilename(title=T["choose"],
                                          filetypes=[("X Archive ZIP", "*.zip")])
        if not path:
            return
        self.hush()
        self._cancel_anim()
        self._show_stage("busy")
        # A multi-gigabyte ZIP took the UI thread down with it before this ran on a
        # worker. Results come back through a Queue because Tk is not thread-safe.
        threading.Thread(target=self._parse_worker, args=(path,), daemon=True).start()
        self.root.after(80, self._poll_parse)

    def _parse_worker(self, path):
        try:
            self._parse_q.put(("ok", archive_parser.analyze_archive(path)))
        except zipfile.BadZipFile:
            self._parse_q.put(("err", "این فایل یک ZIP سالم نیست. همان فایلی را بده که X داده."))
        except Exception as exc:                       # noqa: BLE001 - surfaced to the user
            self._parse_q.put(("err", f"خواندن آرشیو ممکن نشد:\n{exc}"))

    def _poll_parse(self):
        try:
            kind, payload = self._parse_q.get_nowait()
        except pyqueue.Empty:
            self.root.after(80, self._poll_parse)
            return

        if kind == "err":
            self._show_stage("intro")
            self.say(payload)
            return

        self.stats = payload["stats"]
        self.one_way = payload["not_following"]
        self.ignored = payload.get("ignored_files", [])
        self.username = payload.get("account_username", "")
        # Re-read before deciding what belongs in the queue: the web dashboard writes
        # the same file, and anything reviewed there is not new work.
        self.done_ids = set(history_store.load())
        self.queue = [p for p in self.one_way if p["account_id"] not in self.done_ids]
        self.undo_stack = []

        self._show_stage("work")
        self._paint_figures()
        self._paint_tally()
        self._paint_sheet()
        if self.ignored:
            self.say(f'{len(self.ignored)} فایل هم‌نام ولی بی‌ربط نادیده گرفته شد.')

    # --------------------------------------------------------------------------
    # The decisions
    # --------------------------------------------------------------------------
    def _open(self, person):
        try:
            return bool(webbrowser.open_new_tab(person["url"]))
        except Exception:                              # noqa: BLE001 - no browser at all
            return False

    def open_and_record(self):
        if self._busy_card or self._stage_mode != "work" or self.view != "queue":
            return
        if not self.queue:
            return
        person = self.queue[0]
        if not self._open(person):
            self.say("مرورگری برای باز کردن پیدا نشد. آدرس را دستی باز کن.")
            return
        self._record(person)
        self._stamp_and_advance()

    def _record(self, person):
        aid = person["account_id"]
        self.done_ids.add(aid)
        self.undo_stack.append(aid)
        # add() re-reads the file first, so a save from the web app is merged rather
        # than overwritten. The old code wrote its whole in-memory set and lost it.
        history_store.add([aid])

    def _stamp_and_advance(self):
        self._stamp_mark()
        self._busy_card = True

        def lift():
            for item in self.canvas.find_all():
                self.canvas.move(item, 0, -7)

        def finish():
            self._busy_card = False
            self._anim = []
            self._advance(record=True)

        self._anim = [self.root.after(300, lift), self.root.after(460, finish)]

    def _cancel_anim(self):
        """
        Drop a stamp that is still in flight.

        Without this, loading a new archive (or erasing history) while a card
        animates out lets the old timer fire against the new queue and quietly
        retire its first account — a skip nobody asked for and nobody would notice.
        """
        for job in self._anim:
            try:
                self.root.after_cancel(job)
            except tk.TclError:
                pass
        self._anim = []
        self._busy_card = False

    def _advance(self, record):
        if not self.queue:
            return
        person = self.queue.pop(0)
        if not record:
            self.queue.append(person)   # a skip is not progress: it goes to the back
        self._paint_deck()
        self._paint_tally()
        self._paint_figures()

    def skip(self):
        if self._busy_card or self._stage_mode != "work" or self.view != "queue":
            return
        self._advance(record=False)

    def open_batch(self):
        if self._busy_card or self._stage_mode != "work" or not self.queue:
            return
        take = self.queue[:self.batch_size]
        opened = [p for p in take if self._open(p)]
        if not opened:
            self.say("هیچ‌کدام باز نشد. احتمالاً مرورگر پیش‌فرض تنظیم نیست.")
            return
        for person in opened:
            self._record(person)
        # No animation for a batch: a dozen stamps would queue up behind each other
        # and the card you land on would not be the one you expect.
        ids = {p["account_id"] for p in opened}
        self.queue = [p for p in self.queue if p["account_id"] not in ids]
        self._paint_deck()
        self._paint_tally()
        self._paint_figures()
        if len(opened) < len(take):
            self.say(f"{len(take) - len(opened)} پروفایل باز نشد و ثبت هم نشد.")

    def undo(self):
        if not self.undo_stack:
            return
        self._cancel_anim()
        aid = self.undo_stack.pop()
        self.done_ids.discard(aid)
        history_store.remove([aid])
        person = next((p for p in self.one_way if p["account_id"] == aid), None)
        if person:
            self.queue.insert(0, person)
        self._paint_deck()
        self._paint_tally()
        self._paint_figures()
        self._paint_sheet()

    def forget_all(self):
        if not messagebox.askyesno(T["forget"],
                                   "سابقهٔ همهٔ بررسی‌شده‌ها پاک شود؟ این کار برگشت ندارد."):
            return
        history_store.clear()
        self.done_ids = set()
        self.undo_stack = []
        self._cancel_anim()          # a stamp in flight would advance past the refilled queue
        self.queue = list(self.one_way)
        self._paint_deck()
        self._paint_tally()
        self._paint_figures()
        self._paint_sheet()

    def _on_row_open(self, _event=None):
        selected = self.rows.selection()
        if not selected:
            return
        aid = selected[0]
        person = next((p for p in self.one_way if p["account_id"] == aid), None)
        if not person or not self._open(person):
            return
        if self.view == "all" and aid not in self.done_ids:
            self._record(person)
            self.queue = [p for p in self.queue if p["account_id"] != aid]
            self._paint_tally()
            self._paint_figures()
            self._paint_sheet()

    def _on_batch_size(self, _event=None):
        try:
            self.batch_size = int(self.size_box.get())
        except (TypeError, ValueError):
            self.batch_size = 10
        self._paint_tally()

    # --------------------------------------------------------------------------
    # Excel
    # --------------------------------------------------------------------------
    def export_excel(self):
        if not self.one_way:
            self.say("اول یک آرشیو انتخاب کن.")
            return
        # Imported here, not at module scope: openpyxl is only needed if you actually
        # export, and a missing optional dependency should not stop the app opening.
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError:
            self.say("برای خروجی اکسل باید openpyxl نصب باشد:  pip install openpyxl")
            return

        path = filedialog.asksaveasfilename(title=T["excel"], defaultextension=".xlsx",
                                            filetypes=[("Excel", "*.xlsx")])
        if not path:
            return

        pending = [p for p in self.one_way if p["account_id"] not in self.done_ids]
        done = [p for p in self.one_way if p["account_id"] in self.done_ids]

        wb = Workbook()
        head_fill = PatternFill("solid", fgColor="1F4E46")
        head_font = Font(bold=True, color="FCFBF7")

        for index, (title, rows) in enumerate((("مانده در صف", pending),
                                               ("رسیدگی‌شده", done))):
            ws = wb.active if index == 0 else wb.create_sheet()
            ws.title = title
            ws.sheet_view.rightToLeft = True
            ws.append(["یوزرنیم", "آیدی", "نشانی پروفایل"])
            for cell in ws[1]:
                cell.fill, cell.font = head_fill, head_font
                cell.alignment = Alignment(horizontal="center")
            for person in rows:
                handle = person.get("username")
                ws.append([f"@{handle}" if handle else "—", person["account_id"], person["url"]])
            ws.freeze_panes = "A2"
            for column, width in (("A", 26), ("B", 24), ("C", 46)):
                ws.column_dimensions[column].width = width

        try:
            wb.save(path)
        except OSError as exc:
            self.say(f"ذخیره نشد: {exc}")
            return
        self.say(f"ذخیره شد: {path}")


def main():
    root = tk.Tk()
    Desk(root)
    root.mainloop()


if __name__ == "__main__":
    main()
