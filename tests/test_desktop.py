"""
test_desktop.py — drives x_follow_analyzer.py without a display.

faketk stands in for tkinter (see its docstring for why a real shim beats a Mock).
Everything here goes through the app's own public actions — button.invoke(), key
bindings, dialog answers — so a failure means the desktop app is wrong, not that the
test poked at internals.

The history file is redirected to a scratch directory, so running this suite can
never touch the real one.
"""

import io
import json
import os
import sys
import zipfile
from pathlib import Path

# The repo is one level up from tests/; faketk and sandbox live beside this file, which
# is already on sys.path when you run "python tests/test_desktop.py".
sys.path.insert(0, str(Path(__file__).resolve().parent))
import sandbox                                  # noqa: E402
SCRATCH = str(sandbox.redirect("xdesk-"))       # history goes nowhere near the real one

import faketk                                   # noqa: E402
REC = faketk.install()                          # must precede the app import

import history_store                            # noqa: E402
import x_follow_analyzer as xfa                  # noqa: E402

ok = fail = 0


def check(label, cond):
    global ok, fail
    if cond:
        ok += 1
        print(f"  PASS  {label}")
    else:
        fail += 1
        print(f"  FAIL  {label}")


# ---------------------------------------------------------------------------
# A realistic archive: 12 accounts that owe a follow back, 2 mutuals, and the
# lookalike files that used to poison the followers set.
# ---------------------------------------------------------------------------
def ytd(kind, rows):
    return ("﻿" + f"window.YTD.{kind}.part0 = " + json.dumps(rows)).encode("utf-8")


def person(wrapper, aid, handle):
    return {wrapper: {"accountId": aid, "userLink": f"https://twitter.com/{handle}"}}


def build_archive(one_way=12, username="ashka", base="90", handle="user"):
    # base/handle exist so the second archive describes *different* accounts. Sharing
    # IDs between the two fixtures would make the history filter the second load, and
    # every count after that would be measuring the wrong thing.
    following, followers = [], []
    for n in range(1, one_way + 1):
        following.append(person("following", f"{base}{n:02d}", f"{handle}{n:02d}"))
    for aid, who in (("100", "friend_a"), ("101", "friend_b")):
        following.append(person("following", aid, who))
        followers.append(person("follower", aid, who))
    followers.append(person("follower", "300", "fan"))

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as z:
        z.writestr("data/following.js", ytd("following", following))
        z.writestr("data/follower.js", ytd("follower", followers))
        z.writestr("data/account.js",
                   ytd("account", [{"account": {"username": username, "accountId": "1"}}]))
        # Must be ignored: it names accounts we only *requested* to follow.
        z.writestr("data/follower-requests-sent.js",
                   ytd("followerRequestsSent",
                       [person("follower", f"{base}01", f"{handle}01")]))
        z.writestr("data/tweets.js", ytd("tweets", [{"tweet": {"id": "1"}}]))
    return buf.getvalue()


ARCHIVE = os.path.join(SCRATCH, "archive.zip")
with open(ARCHIVE, "wb") as fh:
    fh.write(build_archive())
SECOND = os.path.join(SCRATCH, "second.zip")
with open(SECOND, "wb") as fh:
    fh.write(build_archive(one_way=4, username="other", base="80", handle="other"))
BROKEN = os.path.join(SCRATCH, "broken.zip")
with open(BROKEN, "wb") as fh:
    fh.write(b"this is not a zip file at all")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
opened = []


def fake_open(url):
    opened.append(url)
    return True


xfa.webbrowser.open_new_tab = fake_open


def boot():
    opened.clear()
    root = faketk.Tk()
    return root, xfa.Desk(root)


def card_button(app, text):
    for widget in app.canvas.windows():
        if widget is not None and widget.opts.get("text") == text:
            return widget
    raise AssertionError(f"no card button labelled {text!r}; have "
                         f"{[w.opts.get('text') for w in app.canvas.windows()]}")


def load(root, app, path=ARCHIVE):
    REC.open_path = path
    app.choose_archive()
    root.pump()


def handle_on_card(app):
    for text in app.canvas.texts():
        if text.startswith("@"):
            return text
    return None


# ===========================================================================
print("\n[1] Booting puts a blank docket on the desk")
root, app = boot()
check("the window is titled in Persian", root.title_text == xfa.T["title"])
check("the desk colour is the palette's, not Tk grey",
      root.opts.get("bg") == xfa.PALETTES["light"]["desk"])
check("the intro card invites an archive", xfa.T["intro_lead"] in app.canvas.texts())
check("the counter is hidden until there is something to count", not app.tally.packed)
check("so are the view tabs", not app.views.packed)
check("the batch tray is hidden too", not app.tray.packed)
check("the intro card offers a file button", card_button(app, xfa.T["choose"]) is not None)
check("matplotlib is no longer a dependency", "matplotlib" not in sys.modules)

print("\n[2] Reading the archive happens off the UI thread")
REC.open_path = ARCHIVE
app.choose_archive()
check("the busy card appears immediately", xfa.T["busy_lead"] in app.canvas.texts())
check("and a poll is scheduled rather than blocking", len(root.jobs) == 1)
root.pump()
check("the desk fills once the worker reports back", app.tally.packed and app.views.packed)
check("twelve accounts owe a follow back", len(app.one_way) == 12)
check("the queue starts full", len(app.queue) == 12)
check("the counter shows the queue length", app.lbl_count.opts["text"] == "12")
check("the handle is on the card", handle_on_card(app) == "@user01")
check("the masthead names the account", app.lbl_who.opts["text"] == "@ashka")
check("the ledger counts what you follow", app.figs["following"].opts["text"] == "14")
check("and the mutuals", app.figs["mutual"].opts["text"] == "2")
check("the split bar is drawn now that there are figures", app.split.packed)
check("the request-sent lookalike did not become a follower",
      any(p["account_id"] == "9001" for p in app.one_way))

print("\n[3] Open and record")
card_button(app, xfa.T["open_record"]).invoke()
check("the profile was opened", opened == ["https://x.com/user01"])
check("the stamp is on the card",
      xfa.PALETTES["light"]["stamp"] in app.canvas.fills())
check("input is ignored while the card leaves", app._busy_card)
root.run_pending()
check("the card is released afterwards", not app._busy_card)
check("the queue is one shorter", len(app.queue) == 11)
check("the counter followed it down", app.lbl_count.opts["text"] == "11")
check("the next account is on the card", handle_on_card(app) == "@user02")
check("the decision reached the shared history", history_store.load() == ["9001"])
check("the ledger counts it as cleared", app.figs["cleared"].opts["text"] == "1")

print("\n[4] Next is not progress")
card_button(app, xfa.T["skip"]).invoke()
check("a skip does not shorten the queue", len(app.queue) == 11)
check("the counter does not move", app.lbl_count.opts["text"] == "11")
check("the skipped account moved to the back", handle_on_card(app) == "@user03")
check("and nothing was written to history", history_store.load() == ["9001"])
for _ in range(10):
    app.skip()
check("skipping all the way round returns the same account", handle_on_card(app) == "@user02")

print("\n[5] Batch, with a count you choose")
before = len(app.queue)
app.size_box.set("5")
app.size_box.fire("<<ComboboxSelected>>")
check("the button names the chosen count", "5" in app.btn_batch.opts["text"])
opened.clear()
app.btn_batch.invoke()
check("five profiles opened", len(opened) == 5)
check("five left the queue", len(app.queue) == before - 5)
check("all five reached history", len(history_store.load()) == 6)
check("a batch does not animate", not app._busy_card and not root.jobs)

print("\n[6] Undo puts one back")
top = app.queue[0]["account_id"]
app.undo()
check("the queue grew by one", len(app.queue) == before - 4)
check("history dropped the account", len(history_store.load()) == 5)
check("the restored account is next up", app.queue[0]["account_id"] != top)
check("undo is disabled once the stack is empty",
      [app.undo() for _ in range(9)] and app.btn_undo.opts["state"] == "disabled")

print("\n[7] The reference views")
# [6] emptied the undo stack, so the history is empty again. Clear one account first,
# otherwise the "marked as cleared" assertion would be looking for a mark that is
# correctly absent.
app.open_and_record()
root.run_pending()
app.set_view("all")
check("the sheet replaces the card", app.sheet.packed and not app.canvas.packed)
check("the full list has every account", len(app.rows.get_children()) == 12)
check("cleared rows are marked", any("done" in app.rows.rows[k]["tags"]
                                     for k in app.rows.get_children()))
app.find.insert(0, "user07")
app.find.fire("<KeyRelease>")
check("search narrows the sheet", len(app.rows.get_children()) == 1)
check("and the count reflects it", app.lbl_sheet_count.opts["text"] == "1")
app.find.delete(0, "end")
app.find.fire("<KeyRelease>")
app.set_view("done")
check("the cleared view lists only what you handled",
      len(app.rows.get_children()) == len(history_store.load()))
root.focus = app.find
app._key(app.skip)
check("shortcuts stay quiet while you type in the search box", app.view == "done")
root.focus = None
app.set_view("queue")
check("going back shows the card again", app.canvas.packed and not app.sheet.packed)

print("\n[8] Switching to the dark desk")
app.toggle_theme()
dark = xfa.PALETTES["dark"]
check("the root repainted", root.opts["bg"] == dark["desk"])
check("labels repainted", app.lbl_count.opts["fg"] == dark["ink"])
check("registered buttons repainted", app.btn_batch.opts["bg"] == dark["card"])
check("the treeview style followed",
      app.rows.tags["even"]["background"] == dark["card"])
check("no destroyed card button is left in the theme sweep",
      all(not b.destroyed for b in app._buttons))
check("the card was redrawn in the dark palette", dark["desk"] == app.canvas.opts["bg"])
app.toggle_theme()

print("\n[9] When things fail, the desk says so instead of stopping")
xfa.webbrowser.open_new_tab = lambda url: False
depth = len(app.queue)
app.open_and_record()
check("a profile that will not open is not recorded", "مرورگر" in app.lbl_notice.opts["text"])
check("the notice strip is showing", app.notice.packed)
check("the queue did not move", len(app.queue) == depth)
app.hush()
check("and it can be dismissed", not app.notice.packed)
xfa.webbrowser.open_new_tab = fake_open

load(root, app, BROKEN)
check("a file that is not a ZIP is reported", "ZIP" in app.lbl_notice.opts["text"])
check("and the desk falls back to the intro card",
      xfa.T["intro_lead"] in app.canvas.texts())

print("\n[10] Erasing the history")
# Plant a known history rather than inheriting whatever the sections above left behind:
# this is the assertion that a returning user does not re-review the same accounts.
history_store.save(["9002", "9004", "9006"])
load(root, app, ARCHIVE)
check("the loaded queue skips what history already knows", len(app.queue) == 9)
REC.answer_yes = False
app.forget_all()
check("declining the confirmation changes nothing", len(history_store.load()) == 3)
REC.answer_yes = True
app.forget_all()
check("confirming empties the stored history", history_store.load() == [])
check("everyone comes back to the queue", len(app.queue) == 12)
check("the counter shows the refilled queue", app.lbl_count.opts["text"] == "12")

print("\n[11] A stamp still in flight must not touch the next archive")
# The stamp lands on a timer. Loading a new archive before it fires used to let the old
# timer run against the new queue, retiring an account nobody looked at.
opened.clear()
card_button(app, xfa.T["open_record"]).invoke()
check("a stamp is pending", len(root.jobs) == 2 and app._busy_card)
recorded = history_store.load()
load(root, app, SECOND)
check("the new archive loaded", len(app.one_way) == 4)
check("the stale stamp was cancelled, not fired", not app._busy_card)
check("the new queue is intact", len(app.queue) == 4)
check("its first account is still on the card", handle_on_card(app) == "@other01")
check("only the genuinely recorded account is in history",
      history_store.load() == recorded)

print("\n[12] The history is shared with the web dashboard")
# The old desktop code held a set in memory and wrote it whole, so whichever app
# saved last erased the other's work.
history_store.save(["555000", "555001"])
app.open_and_record()
root.run_pending()
stored = set(history_store.load())
check("a write from elsewhere survives our save", {"555000", "555001"} <= stored)
check("and our own decision is in there too", len(stored) == 3)
load(root, app, SECOND)
check("reloading picks up what the other app recorded",
      len(app.queue) == 3 and len(app.done_ids) == 3)

print("\n[13] Excel export")
target = os.path.join(SCRATCH, "out.xlsx")
REC.save_path = target
app.export_excel()
if "openpyxl" in app.lbl_notice.opts.get("text", ""):
    check("export explains the missing dependency instead of crashing", True)
else:
    check("the workbook was written", os.path.exists(target))
    from openpyxl import load_workbook
    wb = load_workbook(target)
    check("it has a sheet per queue state", wb.sheetnames == ["مانده در صف", "رسیدگی‌شده"])
    check("the pending sheet is right-to-left", wb["مانده در صف"].sheet_view.rightToLeft)
    check("rows match the queue", wb["مانده در صف"].max_row == len(app.queue) + 1)

print("\n[14] Emptying the queue")
guard = 0
while app.queue and guard < 40:
    app.open_and_record()
    root.run_pending()
    guard += 1
check("the queue can be emptied", not app.queue)
check("the empty card explains what to do next", xfa.T["empty_lead"] in app.canvas.texts())
check("the counter reads zero", app.lbl_count.opts["text"] == "0")
check("batch is disabled with nothing to batch", app.btn_batch.opts["state"] == "disabled")
app.skip()
check("pressing next on an empty queue is harmless", app.lbl_count.opts["text"] == "0")
card_button(app, xfa.T["v_done"]).invoke()
check("the empty card offers the cleared list", app.sheet.packed)

print("\n[15] The keyboard drives the whole desk")
load(root, app, ARCHIVE)
app.set_view("queue")
bound = set(root.bindings)
for sequence in ("<Return>", "<space>", "<KeyPress-u>", "<KeyPress-b>",
                 "<Control-o>", "<Control-z>", "<Escape>", "<KeyPress-1>"):
    check(f"{sequence} is bound", sequence in bound)
opened.clear()
before_keys = set(history_store.load())
root.fire("<Return>")
root.run_pending()
check("Enter opens and records",
      len(opened) == 1 and set(history_store.load()) != before_keys)
root.fire("<Control-z>")
check("Ctrl+Z undoes it", set(history_store.load()) == before_keys)

# Every key that is also a printable character has to go through the typing guard. The
# view digits and the theme key were bound around it, so searching for a handle like
# "t3" flipped the theme and jumped to another view instead of typing.
root.focus = app.find
theme_before, view_before = app.theme, app.view
root.fire("<KeyPress-2>")
root.fire("<KeyPress-t>")
check("a digit typed into the search box does not switch view", app.view == view_before)
check("t typed into the search box does not flip the theme", app.theme == theme_before)
root.focus = None
root.fire("<KeyPress-2>")
check("2 switches view when the search box is not focused", app.view == "all")
root.fire("<KeyPress-1>")
check("1 returns to the queue", app.view == "queue")
root.fire("<KeyPress-t>")
check("t flips the theme when the search box is not focused", app.theme != theme_before)
root.fire("<KeyPress-t>")

print(f"\n{'=' * 52}\n  {ok} passed, {fail} failed\n{'=' * 52}")
sys.exit(1 if fail else 0)          # the scratch directory is cleaned up by sandbox
